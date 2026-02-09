"""
Модуль для генерации Excel отчетов
"""
import sys
import os
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from database.crud import AnnouncementCRUD
from typing import Optional


class ExcelReportGenerator:
    """Генератор Excel отчетов"""

    def __init__(self):
        self.reports_dir = os.path.join(
            os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
            'reports'
        )

        # Создать директорию если не существует
        os.makedirs(self.reports_dir, exist_ok=True)

    def _utc_to_local(self, utc_dt: datetime) -> datetime:
        """
        Конвертация UTC времени в местное время Казахстана (UTC+5)

        Args:
            utc_dt: Время в UTC

        Returns:
            Время в часовом поясе Казахстана
        """
        if utc_dt:
            return utc_dt + timedelta(hours=5)
        return None

    def generate_report(
        self,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
        manager_id: Optional[int] = None
    ) -> str:
        """
        Генерировать Excel отчет

        Args:
            start_date: Начальная дата (опционально)
            end_date: Конечная дата (опционально)
            manager_id: ID менеджера (опционально)

        Returns:
            Путь к созданному файлу
        """
        # Получить данные
        announcements = AnnouncementCRUD.get_all_for_report(
            start_date=start_date,
            end_date=end_date,
            manager_id=manager_id
        )

        # Создать workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Отчет по объявлениям"

        # Заголовки
        headers = [
            'Дата создания',
            'Номер объявления',
            'Организация',
            'Юридический адрес',
            'Регион',
            'Лот',
            'Ключевое слово',
            'Менеджер',
            'Статус',
            'Причина отказа',
            'Дата ответа',
            'Детали участия',
            'Ссылка'
        ]

        # Стиль заголовков
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Записать заголовки
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Записать данные
        for row_idx, announcement in enumerate(announcements, start=2):
            # Конвертируем UTC время в местное время Казахстана
            created_at_local = self._utc_to_local(announcement.created_at)
            response_at_local = self._utc_to_local(announcement.response_at)

            ws.cell(row=row_idx, column=1, value=created_at_local.strftime('%Y-%m-%d %H:%M') if created_at_local else '')
            ws.cell(row=row_idx, column=2, value=announcement.announcement_number)
            ws.cell(row=row_idx, column=3, value=announcement.organization_name)
            ws.cell(row=row_idx, column=4, value=announcement.legal_address)
            ws.cell(row=row_idx, column=5, value=announcement.region)
            ws.cell(row=row_idx, column=6, value=announcement.lot_name)
            ws.cell(row=row_idx, column=7, value=announcement.keyword_matched)
            ws.cell(row=row_idx, column=8, value=announcement.manager_name)

            # Статус с цветовой индикацией
            status_cell = ws.cell(row=row_idx, column=9)
            if announcement.status == 'accepted':
                status_cell.value = 'Принято'
                status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif announcement.status == 'rejected':
                status_cell.value = 'Отклонено'
                status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            else:
                status_cell.value = 'Ожидает'
                status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

            ws.cell(row=row_idx, column=10, value=announcement.rejection_reason or '-')

            if response_at_local:
                ws.cell(row=row_idx, column=11, value=response_at_local.strftime('%Y-%m-%d %H:%M'))
            else:
                ws.cell(row=row_idx, column=11, value='-')

            ws.cell(row=row_idx, column=12, value=announcement.participation_details or '-')
            ws.cell(row=row_idx, column=13, value=announcement.announcement_url)

        # Автоподбор ширины столбцов
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            max_length = 0

            for row in range(1, ws.max_row + 1):
                cell = ws[f"{column_letter}{row}"]
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            adjusted_width = min(max_length + 2, 50)  # Максимум 50
            ws.column_dimensions[column_letter].width = adjusted_width

        # Сохранить файл
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"report_{timestamp}.xlsx"
        filepath = os.path.join(self.reports_dir, filename)

        wb.save(filepath)
        print(f"📊 Отчет сохранен: {filepath}")

        return filepath


# Тестирование
if __name__ == '__main__':
    generator = ExcelReportGenerator()
    report_path = generator.generate_report()
    print(f"✅ Отчет создан: {report_path}")
